"""
Business logic for admin operations.
Provides validation and business rules for user/role management.
Also includes hotel guest import functionality.
"""

from models.user import get_all_users, get_user_by_username, get_user_by_email
from models.role import has_users
from datetime import datetime
from typing import Dict, List, Any, Tuple
import re
import openpyxl


def validate_user_creation(username: str, email: str, password: str) -> tuple:
    """
    Validate user creation data.

    Args:
        username: Username to check
        email: Email to check
        password: Password to validate

    Returns:
        Tuple of (is_valid, error_message)
    """
    # Check username exists
    if get_user_by_username(username):
        return False, 'El nombre de usuario ya existe'

    # Validate email format
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        return False, 'Formato de correo electrónico inválido'

    # Check email exists
    if get_user_by_email(email):
        return False, 'El correo electrónico ya existe'

    # Validate password length
    if len(password) < 8:
        return False, 'La contraseña debe tener al menos 8 caracteres'

    # Validate password complexity
    if not any(c.isupper() for c in password):
        return False, 'La contraseña debe contener al menos una letra mayúscula'
    if not any(c.islower() for c in password):
        return False, 'La contraseña debe contener al menos una letra minúscula'
    if not any(c.isdigit() for c in password):
        return False, 'La contraseña debe contener al menos un número'

    return True, ''


def can_delete_user(user_id: int, current_user_id: int) -> tuple:
    """
    Check if user can be deleted.

    Args:
        user_id: User ID to delete
        current_user_id: Current logged-in user ID

    Returns:
        Tuple of (can_delete, error_message)
    """
    # Cannot delete self
    if user_id == current_user_id:
        return False, 'No puede eliminarse a sí mismo'

    # Check if last admin
    from models.user import get_user_by_id
    from models.role import get_role_by_name

    user = get_user_by_id(user_id)
    if not user:
        return False, 'Usuario no encontrado'

    # Get admin role
    admin_role = get_role_by_name('admin')
    if user['role_id'] == admin_role['id']:
        # Count active admins
        all_users = get_all_users(active_only=True)
        admin_count = sum(1 for u in all_users if u['role_id'] == admin_role['id'])

        if admin_count <= 1:
            return False, 'No puede eliminar el último administrador'

    return True, ''


def get_user_activity_summary(user_id: int) -> dict:
    """
    Get activity summary for user.
    Placeholder for future audit log implementation.

    Args:
        user_id: User ID

    Returns:
        Dict with activity stats
    """
    # TODO: Implement in Phase 2 with audit log
    return {
        'total_logins': 0,
        'last_action': None,
        'reservations_created': 0
    }


# ==================== Hotel Guest Import Service ====================

# Column mapping for Spanish headers from PMS Excel exports
COLUMN_MAPPINGS = {
    # Spanish -> English field name
    'nombre': 'first_name',
    'apellidos': 'last_name',
    'apellido': 'last_name',
    'reserva': 'reservation_code',
    'pensión': 'board_type',
    'pension': 'board_type',
    'tipo': 'guest_type',
    'háb.': 'room_number',
    'hab.': 'room_number',
    'habitación': 'room_number',
    'habitacion': 'room_number',
    'room': 'room_number',
    'llegada': 'arrival_date',
    'arrival': 'arrival_date',
    'salida': 'departure_date',
    'departure': 'departure_date',
    'cumpl.': 'birthday',
    'cumpleaños': 'birthday',
    'aniv.': 'anniversary',
    'aniversario': 'anniversary',
    'sexo': 'gender',
    'gender': 'gender',
    'documento': 'document_id',
    'país': 'nationality',
    'pais': 'nationality',
    'country': 'nationality',
    'idioma': 'language',
    'language': 'language',
    'rég fdo.': 'registration_signed',
    'repetidor': 'repeat_guest',
}


def detect_header_row(sheet) -> Tuple[int, Dict[str, int]]:
    """
    Auto-detect the header row and column mappings in an Excel sheet.

    Args:
        sheet: openpyxl worksheet object

    Returns:
        Tuple of (header_row_number, column_mapping_dict)
    """
    # Try first 10 rows to find headers
    for row_num in range(1, 11):
        row_values = [cell.value for cell in sheet[row_num]]

        # Count how many known headers we find
        column_map = {}
        for col_idx, value in enumerate(row_values):
            if value:
                normalized = str(value).lower().strip()
                if normalized in COLUMN_MAPPINGS:
                    field_name = COLUMN_MAPPINGS[normalized]
                    column_map[field_name] = col_idx

        # If we found at least room_number and one date field, this is likely the header
        if 'room_number' in column_map and ('arrival_date' in column_map or 'departure_date' in column_map):
            return row_num, column_map

    # Default to row 3 if detection fails (common in PMS exports)
    return 3, {}


def parse_date(value) -> str:
    """
    Parse date value from Excel cell.

    Args:
        value: Cell value (datetime, string, or None)

    Returns:
        ISO format date string or None
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, str):
        # Try common date formats
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(value.strip(), fmt).date().isoformat()
            except ValueError:
                continue

    return None


def import_hotel_guests_from_excel(
    file_path: str,
    source_name: str = None
) -> Dict[str, Any]:
    """
    Import hotel guests from an Excel file.

    Args:
        file_path: Path to the Excel file
        source_name: Optional source file name for tracking

    Returns:
        Dict with 'created', 'updated', 'errors', 'total' counts
    """
    from models.hotel_guest import upsert_hotel_guest, propagate_room_change

    result = {
        'created': 0,
        'updated': 0,
        'errors': [],
        'total': 0,
        'room_changes': []
    }

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Detect header row and column mapping
        header_row, column_map = detect_header_row(sheet)

        if not column_map:
            # Fallback to manual column mapping for GuestInHouseRpt format
            column_map = {
                'first_name': 0,      # Nombre
                'last_name': 1,       # Apellidos
                'reservation_code': 2, # Reserva
                'board_type': 3,      # Pensión
                'guest_type': 4,      # Tipo
                'room_number': 5,     # Háb.
                'arrival_date': 6,    # Llegada
                'departure_date': 7,  # Salida
                'birthday': 8,        # Cumpl.
                'anniversary': 9,     # Aniv.
                'gender': 10,         # Sexo
                'document_id': 11,    # Documento
                'nationality': 12,    # País
                'language': 13,       # Idioma
                'registration_signed': 14,  # Rég Fdo.
                'repeat_guest': 15,   # Repetidor
            }

        # Process data rows
        for row_num in range(header_row + 1, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_num]]

            # Skip empty rows
            if not any(row_values):
                continue

            result['total'] += 1

            try:
                # Extract values using column map
                def get_value(field: str):
                    if field in column_map:
                        idx = column_map[field]
                        if idx < len(row_values):
                            return row_values[idx]
                    return None

                # Build guest name
                first_name = get_value('first_name') or ''
                last_name = get_value('last_name') or ''
                guest_name = f"{first_name} {last_name}".strip()

                if not guest_name:
                    result['errors'].append(f"Fila {row_num}: Nombre vacío")
                    continue

                # Get room number
                room_number = get_value('room_number')
                if not room_number:
                    result['errors'].append(f"Fila {row_num}: Habitación vacía")
                    continue
                room_number = str(room_number).strip()

                # Parse dates
                arrival_date = parse_date(get_value('arrival_date'))
                departure_date = parse_date(get_value('departure_date'))

                if not arrival_date:
                    result['errors'].append(f"Fila {row_num}: Fecha llegada inválida")
                    continue

                if not departure_date:
                    result['errors'].append(f"Fila {row_num}: Fecha salida inválida")
                    continue

                # Get optional fields
                nationality = get_value('nationality')
                guest_type = get_value('guest_type')
                repeat_count = get_value('repeat_guest')
                booking_reference = get_value('reservation_code')  # "Reserva" column from PMS

                # Determine VIP code based on repeat guest
                vip_code = None
                if repeat_count and str(repeat_count).isdigit() and int(repeat_count) >= 3:
                    vip_code = 'REPEAT'

                # Upsert the guest
                upsert_result = upsert_hotel_guest(
                    room_number=room_number,
                    guest_name=guest_name,
                    arrival_date=arrival_date,
                    departure_date=departure_date,
                    booking_reference=booking_reference,
                    guest_type=guest_type,
                    nationality=nationality,
                    vip_code=vip_code,
                    source_file=source_name
                )

                if upsert_result['action'] == 'created':
                    result['created'] += 1
                else:
                    result['updated'] += 1

                # Handle room change if detected
                if upsert_result.get('room_changed'):
                    propagate_result = propagate_room_change(
                        guest_name=guest_name,
                        old_room=upsert_result['old_room'],
                        new_room=upsert_result['new_room']
                    )

                    result['room_changes'].append({
                        'guest_name': guest_name,
                        'old_room': upsert_result['old_room'],
                        'new_room': upsert_result['new_room'],
                        'customer_updated': propagate_result['customer_updated'],
                        'reservations_updated': propagate_result['reservations_updated']
                    })

            except Exception as e:
                current_app.logger.error(f'Error importing row {row_num}: {e}', exc_info=True)
                result['errors'].append(f"Fila {row_num}: error al procesar datos")

        wb.close()

    except Exception as e:
        current_app.logger.error(f'Error opening import file: {e}', exc_info=True)
        result['errors'].append("Error al abrir archivo")

    return result


def validate_excel_file(file_path: str) -> Tuple[bool, str, Dict[str, Any]]:
    """
    Validate an Excel file before import.

    Args:
        file_path: Path to the Excel file

    Returns:
        Tuple of (is_valid, error_message, preview_data)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Detect header row
        header_row, column_map = detect_header_row(sheet)

        # Check required columns
        if 'room_number' not in column_map:
            return False, 'No se encontró columna de habitación', {}

        if 'arrival_date' not in column_map and 'departure_date' not in column_map:
            return False, 'No se encontraron columnas de fecha', {}

        # Get row count (excluding header and empty rows)
        data_rows = 0
        for row_num in range(header_row + 1, min(sheet.max_row + 1, 1000)):
            row_values = [cell.value for cell in sheet[row_num]]
            if any(row_values):
                data_rows += 1

        # Get preview (first 5 rows)
        preview_rows = []
        for row_num in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_num]]
            if any(row_values):
                preview_rows.append({
                    'row': row_num,
                    'room': str(row_values[column_map.get('room_number', 5)] or ''),
                    'name': f"{row_values[column_map.get('first_name', 0)] or ''} {row_values[column_map.get('last_name', 1)] or ''}".strip(),
                    'arrival': parse_date(row_values[column_map.get('arrival_date', 6)]),
                    'departure': parse_date(row_values[column_map.get('departure_date', 7)]),
                })

        wb.close()

        preview_data = {
            'total_rows': data_rows,
            'columns_found': list(column_map.keys()),
            'preview': preview_rows
        }

        return True, '', preview_data

    except Exception as e:
        current_app.logger.error(f'Error reading Excel file: {e}', exc_info=True)
        return False, "Error al leer archivo. Verifique el formato.", {}
