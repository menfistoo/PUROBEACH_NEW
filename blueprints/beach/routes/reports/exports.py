"""Export routes for reports (customers Excel export)."""
from datetime import date
from flask import request, redirect, url_for, flash, Response
from flask_login import login_required

from utils.decorators import permission_required
from database import get_db
import io


def register_routes(bp):
    """Register export routes on the reports blueprint."""

    @bp.route('/export/customers')
    @login_required
    @permission_required('beach.reports.export')
    def export_customers():
        """Export customers to Excel with professional formatting."""
        return export_customers_handler()


def export_customers_handler() -> Response:
    """
    Generate and return customers Excel export.

    Callable from both the reports blueprint route and the beach
    blueprint convenience route.

    Returns:
        Response: Excel file download response
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Error: openpyxl no esta instalado', 'error')
        return redirect(url_for('beach.customers'))

    # Get filter parameters
    search = request.args.get('search', '')
    customer_type = request.args.get('type', '')
    vip_only = request.args.get('vip', '') == '1'

    # Fetch customers with reservation count and tags
    customers = _get_customers_for_export(
        search=search if search else None,
        customer_type=customer_type if customer_type else None,
        vip_only=vip_only
    )

    # Customer type display mapping
    customer_type_labels = {
        'interno': 'Interno',
        'externo': 'Externo',
    }

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="1A3A5C", end_color="1A3A5C", fill_type="solid"
    )
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = Border(
        left=Side(style='thin', color="D4D4D4"),
        right=Side(style='thin', color="D4D4D4"),
        top=Side(style='thin', color="D4D4D4"),
        bottom=Side(style='thin', color="D4D4D4")
    )
    data_alignment = Alignment(vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill(
        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
    )

    # Title row
    title_font = Font(bold=True, size=14, color="1A3A5C")
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(
        row=1, column=1,
        value="Exportacion de Clientes - PuroBeach"
    )
    title_cell.font = title_font
    title_cell.alignment = Alignment(
        horizontal="center", vertical="center"
    )

    # Subtitle with filter info
    subtitle_parts = []
    if customer_type:
        subtitle_parts.append(
            f"Tipo: {customer_type_labels.get(customer_type, customer_type)}"
        )
    if vip_only:
        subtitle_parts.append("Solo VIP")
    if search:
        subtitle_parts.append(f"Busqueda: {search}")
    subtitle_parts.append(f"Total: {len(customers)} clientes")
    subtitle_text = " | ".join(subtitle_parts)

    ws.merge_cells('A2:G2')
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle_text)
    subtitle_cell.font = Font(size=10, color="666666")
    subtitle_cell.alignment = Alignment(
        horizontal="center", vertical="center"
    )

    # Headers (row 4)
    header_row = 4
    headers = [
        "Nombre", "Tipo", "Habitacion", "Telefono",
        "Email", "Tags", "Total Reservas"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = f'A{header_row + 1}'

    # Data rows
    num_cols = len(headers)
    for row_idx, cust in enumerate(customers, header_row + 1):
        # Nombre
        full_name = cust.get('first_name', '')
        if cust.get('last_name'):
            full_name += f" {cust['last_name']}"
        vip_marker = " [VIP]" if cust.get('vip_status') else ""
        ws.cell(row=row_idx, column=1, value=full_name.strip() + vip_marker)

        # Tipo
        raw_type = cust.get('customer_type', '')
        ws.cell(
            row=row_idx, column=2,
            value=customer_type_labels.get(raw_type, raw_type)
        )

        # Habitacion
        ws.cell(
            row=row_idx, column=3,
            value=cust.get('room_number', '') or '-'
        )

        # Telefono
        phone = cust.get('phone', '') or ''
        country_code = cust.get('country_code', '') or ''
        if phone and country_code:
            phone = f"{country_code} {phone}"
        ws.cell(row=row_idx, column=4, value=phone or '-')

        # Email
        ws.cell(
            row=row_idx, column=5,
            value=cust.get('email', '') or '-'
        )

        # Tags
        ws.cell(
            row=row_idx, column=6,
            value=cust.get('tag_names', '') or '-'
        )

        # Total Reservas
        ws.cell(
            row=row_idx, column=7,
            value=cust.get('reservation_count', 0) or 0
        )

        # Apply styles
        is_alt = (row_idx - header_row) % 2 == 0
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment
            if is_alt:
                cell.fill = alt_fill

        # Center-align specific columns
        for col in [2, 3, 7]:
            ws.cell(row=row_idx, column=col).alignment = center_alignment

    # Auto-adjust column widths
    column_min_widths = {
        1: 22,  # Nombre
        2: 12,  # Tipo
        3: 12,  # Habitacion
        4: 16,  # Telefono
        5: 24,  # Email
        6: 20,  # Tags
        7: 16,  # Total Reservas
    }

    from openpyxl.cell.cell import MergedCell

    for col_cells in ws.columns:
        # Find the first non-merged cell to get column metadata
        anchor_cell = None
        for cell in col_cells:
            if not isinstance(cell, MergedCell):
                anchor_cell = cell
                break
        if anchor_cell is None:
            continue

        col_idx = anchor_cell.column
        max_length = column_min_widths.get(col_idx, 10)
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            try:
                cell_len = len(str(cell.value or ''))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[anchor_cell.column_letter].width = min(
            max_length + 3, 50
        )

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    today = date.today().strftime('%Y-%m-%d')
    filename = f"clientes_{today}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


def _get_customers_for_export(
    search: str = None,
    customer_type: str = None,
    vip_only: bool = False
) -> list:
    """
    Get customers with tags and reservation count for Excel export.

    Args:
        search: Search term for name, phone, email, room
        customer_type: Filter by type ('interno'/'externo')
        vip_only: Only return VIP customers

    Returns:
        list: Customer dicts with tag_names and reservation_count
    """
    with get_db() as conn:
        cursor = conn.cursor()

        query = '''
            SELECT c.*,
                   (SELECT COUNT(*)
                    FROM beach_reservations
                    WHERE customer_id = c.id) as reservation_count,
                   (SELECT GROUP_CONCAT(t.name, ', ')
                    FROM beach_customer_tags ct
                    JOIN beach_tags t ON ct.tag_id = t.id
                    WHERE ct.customer_id = c.id) as tag_names
            FROM beach_customers c
            WHERE 1=1
        '''
        params = []

        if search:
            query += ''' AND (
                c.first_name LIKE ? OR c.last_name LIKE ? OR
                c.email LIKE ? OR c.phone LIKE ? OR c.room_number LIKE ?
            )'''
            search_term = f'%{search}%'
            params.extend([search_term] * 5)

        if customer_type:
            query += ' AND c.customer_type = ?'
            params.append(customer_type)

        if vip_only:
            query += ' AND c.vip_status = 1'

        query += ' ORDER BY c.first_name, c.last_name'

        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]
