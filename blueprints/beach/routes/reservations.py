"""
Beach reservation routes.
Handles reservation list view, quick edit, and Excel export.
Creation is done via the LiveMap.
"""

from flask import current_app, Blueprint, render_template, request, redirect, url_for, flash, Response
from flask_login import login_required, current_user
from utils.decorators import permission_required
from models.reservation import (
    get_reservations_filtered, get_reservation_with_details,
    get_reservation_stats, get_reservation_states,
    delete_reservation, cancel_beach_reservation, get_status_history,
    InvalidStateTransitionError,
)
from datetime import date, datetime
import io

reservations_bp = Blueprint('reservations', __name__)


@reservations_bp.route('/')
@login_required
@permission_required('beach.reservations.view')
def list():
    """Display reservation list with filters."""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    customer_type = request.args.get('type', '')
    state = request.args.get('state', '')
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)

    if not date_from:
        date_from = date.today().strftime('%Y-%m-%d')

    result = get_reservations_filtered(
        date_from=date_from if date_from else None,
        date_to=date_to if date_to else None,
        customer_type=customer_type if customer_type else None,
        state=state if state else None,
        search=search if search else None,
        page=page
    )

    stats = get_reservation_stats(date_from, date_to if date_to else date_from)
    states = get_reservation_states()

    return render_template(
        'beach/reservations.html',
        reservations=result['items'],
        total=result['total'],
        page=result['page'],
        pages=result['pages'],
        stats=stats,
        states=states,
        date_from=date_from,
        date_to=date_to,
        type_filter=customer_type,
        state_filter=state,
        search=search
    )


@reservations_bp.route('/create')
@login_required
@permission_required('beach.reservations.create')
def create():
    """Redirect to map for reservation creation."""
    # Pass any query params to the map (date, customer_id, etc.)
    selected_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    flash('Use el mapa interactivo para crear nuevas reservas', 'info')
    return redirect(url_for('beach.map', date=selected_date))


@reservations_bp.route('/export')
@login_required
@permission_required('beach.reports.export')
def export():
    """Export reservations to Excel with professional formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Error: openpyxl no está instalado', 'error')
        return redirect(url_for('beach.reservations'))

    from models.reservation_queries import get_reservations_for_export

    # Get filter parameters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    customer_type = request.args.get('type', '')
    state_filter = request.args.get('state', '')
    search = request.args.get('search', '')

    if not date_from:
        date_from = date.today().strftime('%Y-%m-%d')

    # Get all reservations for export (includes zone names)
    reservations = get_reservations_for_export(
        date_from=date_from if date_from else None,
        date_to=date_to if date_to else None,
        customer_type=customer_type if customer_type else None,
        state=state_filter if state_filter else None,
        search=search if search else None
    )

    # Payment method display mapping
    payment_method_labels = {
        'efectivo': 'Efectivo',
        'tarjeta': 'Tarjeta',
        'cargo_habitacion': 'Cargo a habitación',
    }

    # Customer type display mapping
    customer_type_labels = {
        'interno': 'Interno',
        'externo': 'Externo',
    }

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color="D4D4D4"),
        right=Side(style='thin', color="D4D4D4"),
        top=Side(style='thin', color="D4D4D4"),
        bottom=Side(style='thin', color="D4D4D4")
    )
    data_alignment = Alignment(vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    currency_format = '#,##0.00 €'
    date_format = 'DD/MM/YYYY'

    # Alternating row fill
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # Title row
    title_font = Font(bold=True, size=14, color="1A3A5C")
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="Exportación de Reservas - PuroBeach")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle with filter info
    subtitle_parts = [f"Desde: {date_from}"]
    if date_to and date_to != date_from:
        subtitle_parts.append(f"Hasta: {date_to}")
    if state_filter:
        subtitle_parts.append(f"Estado: {state_filter}")
    if customer_type:
        subtitle_parts.append(f"Tipo: {customer_type_labels.get(customer_type, customer_type)}")
    subtitle_text = " | ".join(subtitle_parts) + f" | Total: {len(reservations)} reservas"

    ws.merge_cells('A2:K2')
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle_text)
    subtitle_cell.font = Font(size=10, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers (row 4 to leave a gap)
    header_row = 4
    headers = [
        "Ticket", "Cliente", "Tipo Cliente", "Fecha Inicio", "Fecha Fin",
        "Mobiliario", "Zona", "Estado", "Precio", "Método Pago", "Notas"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = f'A{header_row + 1}'

    # Data rows
    num_cols = len(headers)
    for row_idx, res in enumerate(reservations, header_row + 1):
        # Ticket
        ws.cell(row=row_idx, column=1, value=res.get('ticket_number') or f"#{res['id']}")

        # Cliente
        ws.cell(row=row_idx, column=2, value=res.get('customer_name', '').strip())

        # Tipo Cliente
        raw_type = res.get('customer_type', '')
        type_label = customer_type_labels.get(raw_type, raw_type)
        if raw_type == 'interno' and res.get('room_number'):
            type_label += f" (Hab. {res['room_number']})"
        ws.cell(row=row_idx, column=3, value=type_label)

        # Fecha Inicio
        start_val = res.get('reservation_date') or res.get('start_date', '')
        date_cell = ws.cell(row=row_idx, column=4, value=start_val)
        if start_val:
            try:
                date_cell.value = datetime.strptime(start_val, '%Y-%m-%d')
                date_cell.number_format = date_format
            except (ValueError, TypeError):
                date_cell.value = start_val

        # Fecha Fin
        end_val = res.get('end_date', '')
        end_cell = ws.cell(row=row_idx, column=5, value=end_val)
        if end_val:
            try:
                end_cell.value = datetime.strptime(end_val, '%Y-%m-%d')
                end_cell.number_format = date_format
            except (ValueError, TypeError):
                end_cell.value = end_val

        # Mobiliario
        ws.cell(row=row_idx, column=6, value=res.get('furniture_names', '') or '-')

        # Zona
        ws.cell(row=row_idx, column=7, value=res.get('zone_names', '') or '-')

        # Estado
        ws.cell(row=row_idx, column=8, value=res.get('current_state', ''))

        # Precio
        price_cell = ws.cell(row=row_idx, column=9, value=res.get('final_price') or 0)
        price_cell.number_format = currency_format

        # Método Pago
        raw_method = res.get('payment_method', '') or ''
        method_label = payment_method_labels.get(raw_method, raw_method)
        if not method_label and res.get('paid'):
            method_label = 'Pagado'
        elif not method_label:
            method_label = 'Pendiente'
        ws.cell(row=row_idx, column=10, value=method_label)

        # Notas
        ws.cell(row=row_idx, column=11, value=res.get('notes', '') or '')

        # Apply styles to data cells
        is_alt = (row_idx - header_row) % 2 == 0
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment
            if is_alt:
                cell.fill = alt_fill

        # Center-align specific columns
        for col in [1, 3, 4, 5, 8, 10]:
            ws.cell(row=row_idx, column=col).alignment = center_alignment

    # Auto-adjust column widths
    column_min_widths = {
        1: 12,   # Ticket
        2: 20,   # Cliente
        3: 18,   # Tipo Cliente
        4: 14,   # Fecha Inicio
        5: 14,   # Fecha Fin
        6: 18,   # Mobiliario
        7: 15,   # Zona
        8: 14,   # Estado
        9: 12,   # Precio
        10: 18,  # Método Pago
        11: 25,  # Notas
    }

    from openpyxl.cell.cell import MergedCell

    for col_cells in ws.columns:
        # Find the first non-merged cell to get column metadata
        anchor_cell = None
        for cell in col_cells:
            if not isinstance(cell, MergedCell):
                anchor_cell = cell
                break
        if anchor_cell is None:
            continue

        col_idx = anchor_cell.column
        max_length = column_min_widths.get(col_idx, 10)
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            try:
                cell_len = len(str(cell.value or ''))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[anchor_cell.column_letter].width = min(
            max_length + 3, 50
        )

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    filename = f"reservas_{date_from}"
    if date_to and date_to != date_from:
        filename += f"_a_{date_to}"
    filename += ".xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@reservations_bp.route('/<int:reservation_id>')
@login_required
@permission_required('beach.reservations.view')
def detail(reservation_id):
    """Display reservation detail page."""
    reservation = get_reservation_with_details(reservation_id)
    if not reservation:
        flash('Reserva no encontrada', 'error')
        return redirect(url_for('beach.reservations'))

    states = get_reservation_states()
    history = get_status_history(reservation_id)

    return render_template(
        'beach/reservation_unified.html',
        reservation=reservation,
        states=states,
        history=history
    )


@reservations_bp.route('/<int:reservation_id>/edit', methods=['GET'])
@login_required
@permission_required('beach.reservations.edit')
def edit(reservation_id):
    """Redirect to unified detail/edit page."""
    # Unified page handles both view and edit
    return redirect(url_for('beach.reservations_detail', reservation_id=reservation_id))


@reservations_bp.route('/<int:reservation_id>/delete', methods=['POST'])
@login_required
@permission_required('beach.reservations.delete')
def delete(reservation_id):
    """Delete reservation."""
    try:
        delete_reservation(reservation_id)
        flash('Reserva eliminada exitosamente', 'success')
    except Exception as e:
        current_app.logger.error(f'Error: {e}', exc_info=True)
        flash('Error al eliminar. Contacte al administrador.', 'error')

    return redirect(url_for('beach.reservations'))


@reservations_bp.route('/<int:reservation_id>/cancel', methods=['POST'])
@login_required
@permission_required('beach.reservations.change_state')
def cancel(reservation_id):
    """Cancel reservation."""
    reason = request.form.get('reason', '')

    try:
        # Validation is bypassed by default - users can cancel from any state
        cancel_beach_reservation(
            reservation_id,
            cancelled_by=current_user.username if current_user else 'system',
            notes=reason
        )
        flash('Reserva cancelada exitosamente', 'success')
    except InvalidStateTransitionError as e:
        flash(str(e), 'warning')
    except Exception as e:
        current_app.logger.error(f'Error: {e}', exc_info=True)
        flash('Error al cancelar. Contacte al administrador.', 'error')

    return redirect(url_for('beach.reservations_detail', reservation_id=reservation_id))
